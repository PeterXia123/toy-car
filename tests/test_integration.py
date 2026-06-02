from __future__ import annotations

import os
import tempfile

import pandas as pd
import pytest

from eda.checks import data_quality, consistency, trends
from eda.reporting import charts, issue_log


class TestDataQuality:
    def test_missing_values_detected(self, synthetic_df, checks_cfg, variables_cfg):
        findings = data_quality.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        check_ids = [f.check_id for f in findings]
        assert "DQ1_HEATMAP" in check_ids

    def test_negative_values_detected(self, synthetic_df, checks_cfg, variables_cfg):
        findings = data_quality.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        neg_findings = [f for f in findings if f.check_id == "DQ2"]
        assert len(neg_findings) > 0


class TestConsistency:
    def test_default_logic_runs(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        assert len(findings) > 0

    def test_closed_cummax_violation(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        te1 = [f for f in findings if f.check_id == "TE1"]
        assert len(te1) > 0, "Should detect ind_closed cummax violation in ACCT_0001"

    def test_dpd_missing_with_chargeoff(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        df2 = [f for f in findings if f.check_id == "DF2"]
        assert len(df2) > 0, "Should detect DPD missing with ind_CO=1 in ACCT_0002"

    def test_balance_negative(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        lg1 = [f for f in findings if f.check_id == "LG1"]
        assert len(lg1) > 0, "Should detect negative balance in ACCT_0003"

    def test_trend_data_generated(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        df7 = [f for f in findings if f.check_id == "DF7"]
        assert len(df7) > 0
        assert "default_rate" in df7[0].stats

    def test_score_alignment(self, synthetic_df, checks_cfg):
        findings = consistency.run_score_alignment(synthetic_df, checks_cfg, "TEST")
        assert len(findings) > 0

    def test_next_dft_bal_negative(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        nb2 = [f for f in findings if f.check_id == "NB2"]
        assert len(nb2) > 0, "Should detect negative next_dft_bal in ACCT_0004"

    def test_mths_to_dft_negative(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        md2 = [f for f in findings if f.check_id == "MD2"]
        assert len(md2) > 0, "Should detect negative mths_to_dft in ACCT_0006"

    def test_mths_to_dft_mob_inconsistency(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        md4 = [f for f in findings if f.check_id == "MD4"]
        assert len(md4) > 0, "Should detect mths_to_dft vs mob inconsistency in ACCT_0007"


class TestAccountTracking:
    def test_right_censoring_detected(self, synthetic_df, checks_cfg):
        findings = trends.run_account_tracking(synthetic_df, checks_cfg, "TEST")
        at1 = [f for f in findings if f.check_id == "AT1"]
        assert len(at1) > 0, "Should detect right-censored ACCT_0000"



class TestCharts:
    def test_charts_generated(self, synthetic_df, checks_cfg, variables_cfg):
        findings = consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        with tempfile.TemporaryDirectory() as tmpdir:
            charts.generate_all_charts(findings, tmpdir)
            chart_files = [f for f in os.listdir(tmpdir) if f.endswith(".png")]
            assert len(chart_files) > 0, "Should generate at least one chart"


class TestIssueLog:
    def test_issue_log_generated(self, synthetic_df, checks_cfg, variables_cfg):
        findings = data_quality.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
        findings += consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_issue_log.xlsx")
            issue_log.generate(findings, path)
            assert os.path.exists(path)

            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "Issues" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            assert "Examples" in wb.sheetnames

            ws = wb["Issues"]
            assert ws.max_row > 1, "Issues sheet should have data rows"


class TestFullPipeline:
    def test_end_to_end(self, synthetic_df, checks_cfg, variables_cfg):
        with tempfile.TemporaryDirectory() as tmpdir:
            parquet_path = os.path.join(tmpdir, "test_data.parquet")
            synthetic_df.to_parquet(parquet_path)

            findings = []
            findings += data_quality.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
            findings += consistency.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
            findings += consistency.run_score_alignment(synthetic_df, checks_cfg, "TEST")
            findings += trends.run(synthetic_df, checks_cfg, variables_cfg, "TEST")
            findings += trends.run_account_tracking(synthetic_df, checks_cfg, "TEST")

            assert len(findings) > 0

            high = [f for f in findings if f.impact == "High"]
            assert len(high) > 0, "Should have at least one High impact finding"

            charts_dir = os.path.join(tmpdir, "charts")
            os.makedirs(charts_dir)
            charts.generate_all_charts(findings, charts_dir)
            chart_files = os.listdir(charts_dir)
            assert len(chart_files) > 0

            excel_path = os.path.join(tmpdir, "Issue_Log.xlsx")
            issue_log.generate(findings, excel_path)
            assert os.path.exists(excel_path)
