from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from eda.models import Finding, SECTION_ORDER


_FILL_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_MEDIUM = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_NORMAL = Font(size=10)
_FONT_LINK = Font(size=10, color="0563C1", underline="single")
_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")
_FILL_SECTION = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=11)

_HEADERS = [
    "Date",
    "Product",
    "Parameter",
    "Impact",
    "Check ID",
    "Question / Observation",
    "Variable",
    "Downstream Impact",
    "Chart",
    "Comments",
    "EST Comments",
]

_COL_WIDTHS = {
    "Date": 12,
    "Product": 10,
    "Parameter": 18,
    "Impact": 10,
    "Check ID": 10,
    "Question / Observation": 60,
    "Variable": 14,
    "Downstream Impact": 25,
    "Chart": 20,
    "Comments": 30,
    "EST Comments": 30,
}


def generate(findings: list[Finding], output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    issue_findings = [f for f in findings if not f.reference_only]

    wb = openpyxl.Workbook()

    _write_issues_sheet(wb, issue_findings)
    _write_summary_sheet(wb, issue_findings)
    _write_examples_sheet(wb, issue_findings)

    wb.save(output_path)
    return output_path


def _get_section(f: Finding) -> str:
    import re
    m = re.match(r"([A-Z]+)", f.check_id)
    prefix = m.group(1) if m else ""
    return SECTION_ORDER.get(prefix, (99, "Other"))[1]


def _write_issues_sheet(wb: openpyxl.Workbook, findings: list[Finding]) -> None:
    ws = wb.active
    ws.title = "Issues"

    for col_idx, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _BORDER_THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    row_idx = 2
    current_section = None
    for f in findings:
        section = _get_section(f)
        if section != current_section:
            current_section = section
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(_HEADERS))
            cell = ws.cell(row=row_idx, column=1, value=section)
            cell.fill = _FILL_SECTION
            cell.font = _FONT_SECTION
            cell.alignment = Alignment(vertical="center")
            row_idx += 1

        values = [
            f.date,
            f.product,
            f.parameter_str,
            f.impact,
            f.check_id,
            f.question,
            f.variable,
            f.downstream_str,
            "",
            "",
            "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            cell.alignment = _WRAP

        impact_cell = ws.cell(row=row_idx, column=4)
        if f.impact == "High":
            impact_cell.fill = _FILL_HIGH
        elif f.impact == "Medium":
            impact_cell.fill = _FILL_MEDIUM
        elif f.impact == "Low":
            impact_cell.fill = _FILL_LOW

        if f.chart_path:
            chart_cell = ws.cell(row=row_idx, column=9)
            filename = os.path.basename(f.chart_path)
            chart_cell.value = filename
            chart_cell.font = _FONT_LINK
            chart_cell.hyperlink = f.chart_path

        row_idx += 1

    for header_name in _HEADERS:
        col_idx = _HEADERS.index(header_name) + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(header_name, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{row_idx - 1}"


def _write_summary_sheet(wb: openpyxl.Workbook, findings: list[Finding]) -> None:
    ws = wb.create_sheet("Summary")

    parameters = sorted(set(f.parameter_str for f in findings))
    impacts = ["High", "Medium", "Low"]

    counts = {}
    for f in findings:
        p = f.parameter_str
        i = f.impact
        counts[(p, i)] = counts.get((p, i), 0) + 1

    headers = ["Parameter"] + impacts + ["Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _BORDER_THIN

    for row_idx, param in enumerate(parameters, 2):
        ws.cell(row=row_idx, column=1, value=param).border = _BORDER_THIN
        total = 0
        for col_idx, impact in enumerate(impacts, 2):
            c = counts.get((param, impact), 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=c)
            cell.border = _BORDER_THIN
            if impact == "High" and c > 0:
                cell.fill = _FILL_HIGH
            elif impact == "Medium" and c > 0:
                cell.fill = _FILL_MEDIUM
            elif impact == "Low" and c > 0:
                cell.fill = _FILL_LOW
            total += c
        ws.cell(row=row_idx, column=len(impacts) + 2, value=total).border = _BORDER_THIN

    total_row = len(parameters) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = _BORDER_THIN
    for col_idx, impact in enumerate(impacts, 2):
        col_total = sum(counts.get((p, impact), 0) for p in parameters)
        cell = ws.cell(row=total_row, column=col_idx, value=col_total)
        cell.font = Font(bold=True)
        cell.border = _BORDER_THIN
    ws.cell(row=total_row, column=len(impacts) + 2, value=len(findings)).font = Font(bold=True)
    ws.cell(row=total_row, column=len(impacts) + 2).border = _BORDER_THIN

    ws.column_dimensions["A"].width = 25
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    ws.freeze_panes = "A2"


def _write_examples_sheet(wb: openpyxl.Workbook, findings: list[Finding]) -> None:
    ws = wb.create_sheet("Examples")

    row = 1
    for f in findings:
        if f.examples is None or f.examples.empty:
            continue
        if f.impact != "High":
            continue

        cell = ws.cell(row=row, column=1, value=f"[{f.check_id}] {f.variable}: {f.question[:80]}")
        cell.font = Font(bold=True, size=11)
        row += 1

        cols = list(f.examples.columns)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=str(col_name))
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
            cell.border = _BORDER_THIN
        row += 1

        for _, data_row in f.examples.head(20).iterrows():
            for col_idx, col_name in enumerate(cols, 1):
                val = data_row[col_name]
                cell = ws.cell(row=row, column=col_idx, value=_safe_value(val))
                cell.border = _BORDER_THIN
                cell.font = _FONT_NORMAL
            row += 1

        row += 1

    if row == 1:
        ws.cell(row=1, column=1, value="No high-impact findings with example data.")


def _safe_value(val):
    import pandas as pd
    import numpy as np
    if pd.isna(val):
        return ""
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    return val
